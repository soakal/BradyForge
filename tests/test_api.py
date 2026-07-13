import base64
import io
import zipfile
from pathlib import Path

import openpyxl

import bradyforge.config
from bradyforge.api import Api
from bradyforge.settings import default_settings

SAMPLE_MULTI_TAB_PATH = (
    Path(__file__).resolve().parent.parent
    / "samples"
    / "Sample Completed Multi-Tab File.xlsx"
)


def test_get_settings_on_fresh_instance_returns_defaults(tmp_path):
    api = Api(settings_path=tmp_path / "settings.json")

    assert api.get_settings() == default_settings()


def test_save_settings_then_get_settings_round_trips(tmp_path):
    api = Api(settings_path=tmp_path / "settings.json")
    data = {
        "uploads_path": r"\\some-server\share\uploads",
        "label_images_path": r"\\some-server\share\images",
        "fallback_email": "someone@example.com",
    }

    api.save_settings(data)

    assert api.get_settings() == data


def test_save_settings_returns_canonical_persisted_dict(tmp_path):
    api = Api(settings_path=tmp_path / "settings.json")
    data = {
        "uploads_path": r"\\some-server\share\uploads",
        "label_images_path": r"\\some-server\share\images",
        "fallback_email": "someone@example.com",
        "extra_unknown_key": "ignored",
    }

    result = api.save_settings(data)

    assert result == api.get_settings()
    assert "extra_unknown_key" not in result


def test_accept_upload_valid_xlsx_copies_file(tmp_path):
    source_path = tmp_path / "source" / "report.xlsx"
    source_path.parent.mkdir()
    source_path.write_bytes(SAMPLE_MULTI_TAB_PATH.read_bytes())
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()

    api = Api()
    result = api.accept_upload(str(source_path), str(destination_dir))

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert result["filename"] == "report.xlsx"
    assert saved_path == destination_dir / "report.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == source_path.read_bytes()


def test_accept_upload_corrupted_file_returns_error_without_raising(tmp_path):
    source_path = tmp_path / "source" / "not-really.xlsx"
    source_path.parent.mkdir()
    source_path.write_text("this is plain text, not a zip/xlsx file")
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()

    api = Api()
    result = api.accept_upload(str(source_path), str(destination_dir))

    assert result["ok"] is False
    assert "error" in result
    assert list(destination_dir.iterdir()) == []


def test_accept_upload_oversized_file_returns_error_without_raising(
    tmp_path, monkeypatch
):
    monkeypatch.setattr("bradyforge.size_validator.MAX_UPLOAD_SIZE_BYTES", 1000)
    source_path = tmp_path / "source" / "report.xlsx"
    source_path.parent.mkdir()
    source_path.write_bytes(SAMPLE_MULTI_TAB_PATH.read_bytes())
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()

    api = Api()
    result = api.accept_upload(str(source_path), str(destination_dir))

    assert result["ok"] is False
    assert "error" in result
    assert list(destination_dir.iterdir()) == []


def test_accept_upload_filename_collision_gets_timestamp_suffix(tmp_path):
    source_path = tmp_path / "source" / "report.xlsx"
    source_path.parent.mkdir()
    source_path.write_bytes(SAMPLE_MULTI_TAB_PATH.read_bytes())
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    existing_path = destination_dir / "report.xlsx"
    existing_path.write_text("existing upload")

    api = Api()
    result = api.accept_upload(str(source_path), str(destination_dir))

    assert result["ok"] is True
    assert result["filename"] != "report.xlsx"
    assert result["filename"].startswith("report_")
    assert result["filename"].endswith(".xlsx")
    saved_path = Path(result["saved_path"])
    assert saved_path.exists()
    # The pre-existing file must not have been overwritten.
    assert existing_path.read_text() == "existing upload"


def test_accept_upload_unreachable_destination_falls_back_to_local_zip(tmp_path):
    source_path = tmp_path / "source" / "report.xlsx"
    source_path.parent.mkdir()
    source_path.write_bytes(SAMPLE_MULTI_TAB_PATH.read_bytes())

    # A file (not a directory) standing in for the destination, so writing
    # into it as if it were a directory genuinely raises OSError.
    destination_dir = tmp_path / "destination_is_a_file"
    destination_dir.write_text("not a directory")

    fallback_dir = tmp_path / "fallback"
    api = Api(fallback_dir=fallback_dir)
    result = api.accept_upload(str(source_path), str(destination_dir))

    assert result["ok"] is True
    assert result["fallback"] is True
    assert "message" in result

    zip_path = Path(result["zip_path"])
    assert zip_path.exists()

    local_path = Path(result["local_path"])
    assert local_path.exists()
    assert local_path.read_bytes() == source_path.read_bytes()

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        assert len(names) == 1
        assert zf.read(names[0]) == source_path.read_bytes()


def test_accept_upload_without_destination_dir_uses_config_uploads_path(
    tmp_path, monkeypatch
):
    default_destination_dir = tmp_path / "default_destination"
    default_destination_dir.mkdir()
    monkeypatch.setattr(
        bradyforge.config, "UPLOADS_PATH", str(default_destination_dir)
    )
    source_path = tmp_path / "source" / "report.xlsx"
    source_path.parent.mkdir()
    source_path.write_bytes(SAMPLE_MULTI_TAB_PATH.read_bytes())

    api = Api()
    result = api.accept_upload(str(source_path))

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert saved_path == default_destination_dir / "report.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == source_path.read_bytes()


def test_accept_upload_bytes_without_destination_dir_uses_config_uploads_path(
    tmp_path, monkeypatch
):
    default_destination_dir = tmp_path / "default_destination"
    default_destination_dir.mkdir()
    monkeypatch.setattr(
        bradyforge.config, "UPLOADS_PATH", str(default_destination_dir)
    )
    source_bytes = SAMPLE_MULTI_TAB_PATH.read_bytes()
    encoded = base64.b64encode(source_bytes).decode("ascii")

    api = Api()
    result = api.accept_upload_bytes("report.xlsx", encoded)

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert saved_path == default_destination_dir / "report.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == source_bytes


def test_accept_upload_bytes_valid_xlsx_copies_file(tmp_path, monkeypatch):
    fake_temp_dir = tmp_path / "fake_temp"
    monkeypatch.setattr(
        "bradyforge.api.tempfile.mkdtemp", lambda: str(fake_temp_dir.mkdir() or fake_temp_dir)
    )
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    source_bytes = SAMPLE_MULTI_TAB_PATH.read_bytes()
    encoded = base64.b64encode(source_bytes).decode("ascii")

    api = Api()
    result = api.accept_upload_bytes("report.xlsx", encoded, str(destination_dir))

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert result["filename"] == "report.xlsx"
    assert saved_path == destination_dir / "report.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == source_bytes
    assert not fake_temp_dir.exists()


def test_accept_upload_bytes_data_url_prefix_is_stripped(tmp_path):
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    source_bytes = SAMPLE_MULTI_TAB_PATH.read_bytes()
    encoded = base64.b64encode(source_bytes).decode("ascii")
    data_url = (
        "data:application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet;base64," + encoded
    )

    api = Api()
    result = api.accept_upload_bytes("report.xlsx", data_url, str(destination_dir))

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert result["filename"] == "report.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == source_bytes


def test_accept_upload_bytes_corrupted_content_returns_error_without_raising(
    tmp_path, monkeypatch
):
    fake_temp_dir = tmp_path / "fake_temp"
    monkeypatch.setattr(
        "bradyforge.api.tempfile.mkdtemp", lambda: str(fake_temp_dir.mkdir() or fake_temp_dir)
    )
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    encoded = base64.b64encode(b"this is plain text, not a zip/xlsx file").decode(
        "ascii"
    )

    api = Api()
    result = api.accept_upload_bytes("not-really.xlsx", encoded, str(destination_dir))

    assert result["ok"] is False
    assert "error" in result
    assert list(destination_dir.iterdir()) == []
    assert not fake_temp_dir.exists()


def test_submit_generic_labels_valid_rows_writes_workbook(tmp_path):
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    rows = [
        {"line1": "Widget A", "line2": "Part 123", "line3": "Rev 1", "qty": 10},
        {"line1": "Widget B", "line2": "Part 456", "line3": "Rev 2", "qty": 5},
    ]

    api = Api()
    result = api.submit_generic_labels(rows, "labels.xlsx", str(destination_dir))

    assert result["ok"] is True
    assert result["filename"] == "labels.xlsx"
    saved_path = Path(result["saved_path"])
    assert saved_path == destination_dir / "labels.xlsx"
    assert saved_path.exists()

    workbook = openpyxl.load_workbook(saved_path)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))
    assert all_rows[0] == ("Line 1", "Line2", "Line3", "qty")
    assert all_rows[1:] == [
        (row["line1"], row["line2"], row["line3"], row["qty"]) for row in rows
    ]


def test_submit_generic_labels_empty_rows_returns_error_without_writing(tmp_path):
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()

    api = Api()
    result = api.submit_generic_labels([], "labels.xlsx", str(destination_dir))

    assert result["ok"] is False
    assert "error" in result
    assert list(destination_dir.iterdir()) == []


def test_submit_generic_labels_unreachable_destination_falls_back_to_local_zip(
    tmp_path,
):
    # A file (not a directory) standing in for the destination, so writing
    # into it as if it were a directory genuinely raises OSError.
    destination_dir = tmp_path / "destination_is_a_file"
    destination_dir.write_text("not a directory")

    fallback_dir = tmp_path / "fallback"
    rows = [
        {"line1": "Widget A", "line2": "Part 123", "line3": "Rev 1", "qty": 10},
        {"line1": "Widget B", "line2": "Part 456", "line3": "Rev 2", "qty": 5},
    ]

    api = Api(fallback_dir=fallback_dir)
    result = api.submit_generic_labels(rows, "labels.xlsx", str(destination_dir))

    assert result["ok"] is True
    assert result["fallback"] is True
    assert "message" in result

    zip_path = Path(result["zip_path"])
    assert zip_path.exists()

    local_path = Path(result["local_path"])
    assert local_path.exists()

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        assert len(names) == 1
        assert names[0] == "labels.xlsx"
        workbook_bytes = zf.read(names[0])

    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))
    assert all_rows[0] == ("Line 1", "Line2", "Line3", "qty")
    assert all_rows[1:] == [
        (row["line1"], row["line2"], row["line3"], row["qty"]) for row in rows
    ]


def test_submit_generic_labels_without_destination_dir_uses_config_uploads_path(
    tmp_path, monkeypatch
):
    default_destination_dir = tmp_path / "default_destination"
    default_destination_dir.mkdir()
    monkeypatch.setattr(
        bradyforge.config, "UPLOADS_PATH", str(default_destination_dir)
    )
    rows = [{"line1": "Widget A", "line2": "Part 123", "line3": "Rev 1", "qty": 10}]

    api = Api()
    result = api.submit_generic_labels(rows, "labels.xlsx")

    assert result["ok"] is True
    saved_path = Path(result["saved_path"])
    assert saved_path == default_destination_dir / "labels.xlsx"
    assert saved_path.exists()

    workbook = openpyxl.load_workbook(saved_path)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))
    assert all_rows[0] == ("Line 1", "Line2", "Line3", "qty")
    assert all_rows[1:] == [
        (row["line1"], row["line2"], row["line3"], row["qty"]) for row in rows
    ]


def test_submit_generic_labels_filename_collision_gets_timestamp_suffix(tmp_path):
    destination_dir = tmp_path / "destination"
    destination_dir.mkdir()
    existing_path = destination_dir / "labels.xlsx"
    existing_path.write_text("existing labels")
    rows = [{"line1": "Widget A", "line2": "Part 123", "line3": "Rev 1", "qty": 10}]

    api = Api()
    result = api.submit_generic_labels(rows, "labels.xlsx", str(destination_dir))

    assert result["ok"] is True
    assert result["filename"] != "labels.xlsx"
    assert result["filename"].startswith("labels_")
    assert result["filename"].endswith(".xlsx")
    saved_path = Path(result["saved_path"])
    assert saved_path.exists()
    # The pre-existing file must not have been overwritten.
    assert existing_path.read_text() == "existing labels"


def test_get_label_images_with_explicit_dir_returns_data_urls(tmp_path):
    images_dir = tmp_path / "images"
    images_dir.mkdir()
    png_bytes = b"fake-png-bytes-123"
    jpg_bytes = b"fake-jpg-bytes-456"
    gif_bytes = b"fake-gif-bytes-789"
    (images_dir / "b.png").write_bytes(png_bytes)
    (images_dir / "a.jpg").write_bytes(jpg_bytes)
    (images_dir / "c.gif").write_bytes(gif_bytes)

    api = Api()
    result = api.get_label_images(str(images_dir))

    assert [entry["filename"] for entry in result] == ["a.jpg", "b.png", "c.gif"]

    by_filename = {entry["filename"]: entry["data_url"] for entry in result}

    assert by_filename["a.jpg"].startswith("data:image/jpeg;base64,")
    assert by_filename["b.png"].startswith("data:image/png;base64,")
    assert by_filename["c.gif"].startswith("data:image/gif;base64,")

    for filename, expected_bytes in (
        ("a.jpg", jpg_bytes),
        ("b.png", png_bytes),
        ("c.gif", gif_bytes),
    ):
        _, _, encoded = by_filename[filename].partition("base64,")
        assert base64.b64decode(encoded) == expected_bytes


def test_get_label_images_without_images_dir_uses_config_label_images_path(
    tmp_path, monkeypatch
):
    default_images_dir = tmp_path / "default_images"
    default_images_dir.mkdir()
    monkeypatch.setattr(
        bradyforge.config, "LABEL_IMAGES_PATH", str(default_images_dir)
    )
    image_bytes = b"fake-bmp-bytes"
    (default_images_dir / "logo.bmp").write_bytes(image_bytes)

    api = Api()
    result = api.get_label_images()

    assert len(result) == 1
    assert result[0]["filename"] == "logo.bmp"
    assert result[0]["data_url"].startswith("data:image/bmp;base64,")
    _, _, encoded = result[0]["data_url"].partition("base64,")
    assert base64.b64decode(encoded) == image_bytes


def test_get_label_images_empty_directory_returns_empty_list(tmp_path):
    images_dir = tmp_path / "images"
    images_dir.mkdir()

    api = Api()
    result = api.get_label_images(str(images_dir))

    assert result == []


def test_get_label_images_skips_unreadable_file_without_raising(tmp_path, monkeypatch):
    images_dir = tmp_path / "images"
    images_dir.mkdir()
    good_bytes = b"good-png-bytes"
    (images_dir / "good.png").write_bytes(good_bytes)
    (images_dir / "bad.png").write_bytes(b"bad-png-bytes")

    real_open = open

    def flaky_open(file, *args, **kwargs):
        if str(file).endswith("bad.png"):
            raise OSError("simulated unreadable file")
        return real_open(file, *args, **kwargs)

    monkeypatch.setattr("builtins.open", flaky_open)

    api = Api()
    result = api.get_label_images(str(images_dir))

    assert [entry["filename"] for entry in result] == ["good.png"]
    _, _, encoded = result[0]["data_url"].partition("base64,")
    assert base64.b64decode(encoded) == good_bytes
