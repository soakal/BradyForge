from pathlib import Path

import openpyxl

from bradyforge.generic_writer import write_generic_workbook

SAMPLE_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "samples" / "Generic Label template.xlsx"
)


def _sample_header():
    workbook = openpyxl.load_workbook(SAMPLE_TEMPLATE_PATH)
    sheet = workbook.active
    return next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))


def test_header_matches_sample_template(tmp_path):
    dest_path = tmp_path / "output.xlsx"
    write_generic_workbook([], dest_path)

    workbook = openpyxl.load_workbook(dest_path)
    sheet = workbook.active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    assert header == _sample_header()


def test_rows_round_trip_in_order(tmp_path):
    dest_path = tmp_path / "output.xlsx"
    rows = [
        {"line1": "Widget A", "line2": "Part 123", "line3": "Rev 1", "qty": 10},
        {"line1": "Widget B", "line2": "Part 456", "line3": "Rev 2", "qty": 5},
        {"line1": "Widget C", "line2": "Part 789", "line3": "Rev 3", "qty": 1},
    ]
    write_generic_workbook(rows, dest_path)

    workbook = openpyxl.load_workbook(dest_path)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))

    assert all_rows[0] == _sample_header()
    assert all_rows[1:] == [
        (row["line1"], row["line2"], row["line3"], row["qty"]) for row in rows
    ]
